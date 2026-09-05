"""Permit file generation: XLSX (conditional colours + summary sheet) and CSV (UTF-8 BOM for Excel)."""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app.audit.log import record
from app.core.security import mask_id
from app.db.models import Batch, Document
from app.db.session import session_scope
from app.engines.rules import RulesRepository
from app.services.store import current_decision, decision_to_dict, load_extraction

COLUMNS = [
    ("File Name", "file_name"), ("Iqama No.", "iqama_no"), ("Name (AR)", "name_ar"), ("Name (EN)", "name_en"),
    ("Nationality", "nationality"), ("Occupation (AR)", "occupation"), ("Occupation (EN)", "occupation_en"),
    ("Occupation Code", "occupation_code"), ("Employer", "employer_name"), ("Employer ID", "employer_id"),
    ("Employer Type", "employer_type"), ("Issue Place", "issue_place"), ("Expiry Date", "expiry_date"),
    ("Days Remaining", "days_remaining"), ("Iqama Status", "iqama_status"), ("Nationality Status", "nationality_status"),
    ("Occupation Status", "occupation_status"), ("Employer Status", "employer_status"), ("Final Decision", "final_decision"),
    ("Rejection Reasons", "rejection_reasons"), ("Review Triggers", "review_triggers"), ("OCR Confidence (min)", "conf_min"),
    ("OCR Confidence (avg)", "conf_avg"), ("Image Quality", "quality"), ("Reviewed By", "reviewed_by"),
    ("Reviewed At", "reviewed_at"), ("Card Layout", "layout"), ("Note", "note"), ("Rules Version", "rules_version"),
]
FILL = {"APPROVED": "C6EFCE", "REJECTED": "FFC7CE", "MANUAL_REVIEW": "FFEB9C"}
DISCLAIMER = ("This screening is based on the printed card image only, not on official government systems. "
              "It does not detect forgery and does not replace verification with the competent authorities.")


class ReportService:
    def __init__(self, rules: RulesRepository) -> None:
        self.rules = rules

    def rows(self, batch_id: int, unmask: bool = False) -> list[dict]:
        out = []
        with session_scope() as s:
            b = s.get(Batch, batch_id)
            if b is None:
                raise KeyError(f"batch {batch_id} not found")
            for d in sorted(b.documents, key=lambda d: d.id):
                x = load_extraction(s, d)
                dec = decision_to_dict(current_decision(d))
                checks = {c["check"]: c for c in (dec or {}).get("checks", [])}
                confs = [f.confidence for f in x.fields.values() if f.field in self.rules.active.config.ocr.critical_fields]
                occ = checks.get("OCCUPATION", {}).get("details", {})
                review = max(d.reviews, key=lambda r: r.id) if d.reviews else None
                iq = x.value("iqama_no")
                emp = x.value("employer_id")
                out.append({
                    "file_name": d.original_filename + (f" (p{d.page_no})" if d.page_no > 1 else ""),
                    "iqama_no": iq if unmask else mask_id(iq),
                    "name_ar": x.value("name_ar"), "name_en": x.value("name_en"),
                    "nationality": x.value("nationality"),
                    "occupation": x.value("occupation"), "occupation_en": occ.get("matched_en"), "occupation_code": None,
                    "employer_name": x.value("employer_name"), "employer_id": emp if unmask else mask_id(emp),
                    "employer_type": checks.get("EMPLOYER", {}).get("label"),
                    "issue_place": x.value("issue_place"),
                    "expiry_date": x.value("expiry_date"),
                    "days_remaining": checks.get("EXPIRY", {}).get("details", {}).get("days_remaining"),
                    "iqama_status": checks.get("EXPIRY", {}).get("label"),
                    "nationality_status": checks.get("NATIONALITY", {}).get("label"),
                    "occupation_status": checks.get("OCCUPATION", {}).get("label"),
                    "employer_status": checks.get("EMPLOYER", {}).get("label"),
                    "final_decision": (dec or {}).get("status") or d.status,
                    "rejection_reasons": "; ".join(f"{i+1}. {r}" for i, r in enumerate((dec or {}).get("reasons", []))),
                    "review_triggers": "; ".join((dec or {}).get("review_triggers", [])),
                    "conf_min": round(min(confs), 2) if confs else None,
                    "conf_avg": round(sum(confs) / len(confs), 2) if confs else None,
                    "quality": d.quality_score,
                    "reviewed_by": review.reviewer if review else None,
                    "reviewed_at": review.submitted_at.isoformat() if review and review.submitted_at else None,
                    "rules_version": (dec or {}).get("rules_version"),
                    "layout": x.layout,
                    "note": next((t.split(": ", 1)[1] for t in (dec or {}).get("review_triggers", []) if t.startswith("OLD_LAYOUT")), None),
                    "error": d.error_msg,
                })
        return out

    def summary(self, batch_id: int) -> dict:
        rows = self.rows(batch_id)
        c = {"total": len(rows), "approved": 0, "rejected": 0, "manual_review": 0, "expired": 0,
             "individual_employer": 0, "excluded_occupation": 0, "nationality_not_approved": 0, "errors": 0}
        for r in rows:
            fd = r["final_decision"]
            if fd == "APPROVED": c["approved"] += 1
            elif fd == "REJECTED": c["rejected"] += 1
            elif fd == "MANUAL_REVIEW": c["manual_review"] += 1
            if r["error"]: c["errors"] += 1
            if r["iqama_status"] == "EXPIRED": c["expired"] += 1
            if r["employer_status"] == "INDIVIDUAL": c["individual_employer"] += 1
            if r["occupation_status"] in ("EXCLUDED", "EXCLUDED_FUZZY"): c["excluded_occupation"] += 1
            if r["nationality_status"] == "NOT_APPROVED": c["nationality_not_approved"] += 1
        return c

    def to_csv(self, batch_id: int, unmask: bool, actor: str) -> bytes:
        rows = self.rows(batch_id, unmask)
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([h for h, _ in COLUMNS])
        for r in rows:
            w.writerow([r.get(k) for _, k in COLUMNS])
        w.writerow([]); w.writerow([DISCLAIMER])
        self._audit(batch_id, "csv", unmask, actor, len(rows))
        return ("﻿" + buf.getvalue()).encode("utf-8")

    def to_xlsx(self, batch_id: int, unmask: bool, actor: str) -> bytes:
        rows = self.rows(batch_id, unmask)
        summ = self.summary(batch_id)
        wb = Workbook()
        ws = wb.active; ws.title = "Permit File"
        ws.sheet_view.rightToLeft = False
        bold = Font(bold=True)
        for j, (h, _) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=j, value=h); c.font = bold; c.alignment = Alignment(horizontal="center")
        for i, r in enumerate(rows, start=2):
            for j, (_, k) in enumerate(COLUMNS, start=1):
                ws.cell(row=i, column=j, value=r.get(k))
            fill = FILL.get(r["final_decision"])
            if fill:
                ws.cell(row=i, column=[k for _, k in COLUMNS].index("final_decision") + 1).fill = PatternFill("solid", fgColor=fill)
        for j, (h, k) in enumerate(COLUMNS, start=1):
            width = max([len(str(h))] + [len(str(r.get(k) or "")) for r in rows]) + 2
            ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, width))
        ws.freeze_panes = "A2"
        s2 = wb.create_sheet("Summary")
        s2.append(["Generated at", datetime.now(timezone.utc).isoformat()])
        s2.append(["Generated by", actor])
        s2.append(["Rules version", self.rules.active.version])
        s2.append(["Rules note", self.rules.active.config.version_note])
        s2.append([])
        for k, v in summ.items():
            s2.append([k.replace("_", " ").title(), v])
        s2.append([]); s2.append(["Disclaimer", DISCLAIMER])
        s2.column_dimensions["A"].width = 28; s2.column_dimensions["B"].width = 60
        out = io.BytesIO(); wb.save(out)
        self._audit(batch_id, "xlsx", unmask, actor, len(rows))
        return out.getvalue()

    def _audit(self, batch_id: int, fmt: str, unmask: bool, actor: str, n: int) -> None:
        with session_scope() as s:
            record(s, actor, "REPORT_EXPORTED", "batch", batch_id, {"format": fmt, "unmasked": unmask, "rows": n})
